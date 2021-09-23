import pymssql
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.types import VARCHAR, Float, Integer, Date, Numeric
import win32com.client
from openpyxl  import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import time
# conn = pymssql.connect(server="192.168.10.9", user="sj_xudq", password="oQJhkk#53", database="db_danqi")
# engine = create_engine('mssql+pymssql://sj_xudq:oQJhkk#53@192.168.10.9/db_danqi')
# # cur = conn.cursor()
# # cur.execute("exec db_danqi.dbo.M3_360_p")
# # cur.close()
# time.sleep(60)
# detail=pd.read_sql("select * from db_danqi.dbo.M3_360_1",con=engine)
# print(detail)
# num = detail.shape[0]
# conn.close()
#
# wb=load_workbook(r"C:\Users\Administrator\Desktop\360-M3-模板-月末.xlsx")
# wb.remove(wb['Sheet1'])
# wb.create_sheet('Sheet1',index=None)
#
# for row in dataframe_to_rows(detail,index=False):
# 	wb['Sheet1'].append(row)
# wb.save(r"C:\Users\Administrator\Desktop\360-M3-模板-月末.xlsx")
#
# def just_open(filename):
#     xlApp = win32com.client.Dispatch("Excel.Application")   #Dispatch("Excel.Application")
#     xlApp.Visible = False
#     xlBook = xlApp.Workbooks.Open(filename)
#     xlBook.Save()
#     xlBook.Close()
# just_open(r"C:\Users\Administrator\Desktop\360-M3-模板-月末.xlsx")

import pandas as pd
df=pd.read_excel(r"C:\Users\Administrator\Desktop\360-M3-模板-月末.xlsx",sheet_name='案件导入模板')
# num 更改
df=df[0:119]
df['卡号']=df['卡号'].astype('str')
df['证件号']=df['证件号'].astype('str')


df1=pd.read_excel(r"C:\Users\Administrator\Desktop\QHJR-20210901-M3.xls")
df1['个案序列号'][df1['个案序列号'].notnull()]=df1['个案序列号'][df1['个案序列号'].notnull()].str.replace('.*','',regex=True)
# print(df1)
df2=pd.concat([df1,df],axis=0)
df2['卡号']=df2['卡号'].astype('str')
df2['证件号']=df2['证件号'].astype('str')
df2.to_excel(r"C:\Users\Administrator\Desktop\QHJR-20210901-M3.xls",index=None)


# from openpyxl import load_workbook
# wb = load_workbook(r"C:\Users\Administrator\Desktop\360.xlsx",data_only=True)
# ws = wb['案件导入模板']
# rows=ws.max_row  #获取行数
# cols=ws.max_column  #获取列数
# print(rows,cols)
# print(ws['A'])
# for i in ws.values:
#     print(i)

#遍历行
# for i in ws.values:
#     print(i)

# wb.get_sheet_by_name('Sheet1')
# 实例化
# wb = Workbook()
# 激活 worksheet
# ws = wb.active


# df=pd.read_excel(r"C:\Users\Administrator\Desktop\360-M3-模板-月末.xls",sheet_name=None)
# sheet_name=list(df.keys())
#
# df_1=df['Sheet1']
# df_11=df_1.drop(df_1.index[0:df_1.shape[0]])
# detail=pd.read_sql("select * from db_danqi.dbo.M3_360",con=engine)
# df['Sheet1']=df_11.append(detail)
# with pd.ExcelWriter(r"C:\Users\Administrator\Desktop\360-M3-模板-月末.xls") as writer:
#     df['Sheet1'].to_excel(writer,index=None,sheet_name='Sheet1')


